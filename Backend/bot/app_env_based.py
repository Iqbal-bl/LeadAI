# routers/schema_router.py
import os
import re
import csv
from pathlib import Path
from typing import Dict, List, Optional, Union

from fastapi import APIRouter, HTTPException, Query, Body
from openpyxl import load_workbook
from threading import Lock

from bot.claim_data_get import ClaimDataPrompt  # your class
from bot.synonym_manager import SynonymManager
from Domain import models, schema

router = APIRouter(prefix="/schema", tags=["schema"])

# ---------------------- Config ----------------------
FILE_PATH = os.getenv("FILE_PATH", "").strip()
if not FILE_PATH:
    raise RuntimeError("FILE_PATH env var not set.")
FILE_PATH = str(Path(FILE_PATH).resolve())

DEFAULT_SHEET = os.getenv("SHEET", "").strip() or None  # Excel only
EXT = Path(FILE_PATH).suffix.lower()
if EXT not in [".csv", ".xlsx"]:
    raise RuntimeError("Only .csv or .xlsx are supported via FILE_PATH.")
KIND = "csv" if EXT == ".csv" else "excel"

# single-process write lock to avoid races
_file_lock = Lock()

# ---------------------- Alias utils ----------------------
_SPLIT = re.compile(r"[\/,;|]+")

def _norm_token(t: str) -> str:
    t = t.replace("\u00A0", " ")
    t = t.replace("_", " ")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t

def _store_token(t: str) -> str:
    return t.replace(" ", "_")

def parse_aliases_from_header(header_text: str) -> List[str]:
    s = str(header_text or "").strip()
    m = re.match(r"^(.*?)(?:\((.*?)\))?$", s)
    base = (m.group(1) or "").strip()
    inside = (m.group(2) or "").strip()

    tokens: List[str] = []
    if base:
        tokens += _SPLIT.split(base)
    if inside:
        tokens += _SPLIT.split(inside)

    out: List[str] = []
    for tok in tokens:
        n = _norm_token(tok)
        if n and n not in out:
            out.append(n)
    return out

def first_alias_of(header_text: str) -> str:
    aliases = parse_aliases_from_header(header_text)
    return aliases[0] if aliases else _norm_token(header_text)

def first_alias_storage(header_text: str) -> str:
    s = str(header_text or "").strip()
    base = s.split("(", 1)[0].strip()
    first = _SPLIT.split(base)[0].strip() if base else ""
    return first or s

def render_aliases_with_base(base_storage: str, rest_aliases_norm: List[str]) -> str:
    base_norm = _norm_token(base_storage)
    out_parts: List[str] = [base_storage] if base_storage else []
    for a in rest_aliases_norm:
        if a and a != base_norm:
            out_parts.append(_store_token(a))
    return "/".join(out_parts)

def _compose_new_header(current_header: str, new_aliases: List[str], mode: str) -> str:
    curr_norm = parse_aliases_from_header(current_header)
    base_storage = first_alias_storage(current_header)
    provided_norm = [_norm_token(a) for a in new_aliases if _norm_token(a)]

    if mode == "merge":
        merged = curr_norm[:]
        for a in provided_norm:
            if a not in merged:
                merged.append(a)
        return render_aliases_with_base(base_storage, merged)

    return render_aliases_with_base(base_storage, provided_norm)

# ---------------------- CSV helpers ----------------------
def detect_csv_headers(file_path: str) -> List[str]:
    p = Path(file_path)
    if not p.exists():
        raise HTTPException(status_code=404, detail="FILE_PATH not found.")
    with p.open("r", newline="", encoding="utf-8") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        headers = next(reader, [])
        if not headers:
            raise HTTPException(status_code=400, detail="No header row found in CSV.")
        return headers

def apply_synonyms_to_csv(file_path: str, syn_map_array: Dict[str, List[str]], mode: str) -> List[str]:
    p = Path(file_path)
    with p.open("r", newline="", encoding="utf-8") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        rows = list(csv.reader(f, dialect))

    if not rows:
        raise HTTPException(status_code=400, detail="CSV appears empty.")

    original_header = rows[0]
    norm_keys: Dict[str, List[str]] = { _norm_token(k): v for k, v in syn_map_array.items() }

    new_header: List[str] = []
    for h in original_header:
        first_key = first_alias_of(h)
        raw_key = _norm_token(h)
        payload = None
        if first_key in norm_keys:
            payload = norm_keys[first_key]
        elif raw_key in norm_keys:
            payload = norm_keys[raw_key]

        if payload is not None:
            new_header.append(_compose_new_header(h, payload, mode))
        else:
            new_header.append(h)

    rows[0] = new_header
    with p.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, dialect)
        writer.writerows(rows)

    return new_header

# ---------------------- Excel helpers ----------------------
def detect_excel_headers(file_path: str, sheet: Optional[str] = None) -> Dict[str, List[str]]:
    wb = load_workbook(filename=file_path, data_only=False)
    sheets = [sheet] if sheet else wb.sheetnames
    out: Dict[str, List[str]] = {}
    for sn in sheets:
        ws = wb[sn]
        row1 = ws[1]
        headers: List[str] = []
        for cell in row1:
            headers.append("" if cell.value is None else str(cell.value))
        out[sn] = headers
    return out

def apply_synonyms_to_excel(file_path: str, syn_map_array: Dict[str, List[str]], sheet: Optional[str], mode: str) -> Dict[str, List[str]]:
    wb = load_workbook(filename=file_path)
    targets = [sheet] if sheet else wb.sheetnames
    norm_keys: Dict[str, List[str]] = { _norm_token(k): v for k, v in syn_map_array.items() }

    for sn in targets:
        ws = wb[sn]
        for cell in ws[1]:
            current = "" if cell.value is None else str(cell.value)
            first_key = first_alias_of(current)
            raw_key = _norm_token(current)

            payload = None
            if first_key in norm_keys:
                payload = norm_keys[first_key]
            elif raw_key in norm_keys:
                payload = norm_keys[raw_key]

            if payload is not None:
                cell.value = _compose_new_header(current, payload, mode)

    wb.save(filename=file_path)
    return detect_excel_headers(file_path, sheet=None if sheet is None else sheet)

# ---------------------- Routes ----------------------
@router.get("/info", response_model=schema.SchemaInfoResponse, summary="Get Schema Info", description="**Output:** JSON with file path, kind (csv/excel), and default sheet.<br>**Input:** None")
def info():
    return {"file_path": FILE_PATH, "kind": KIND, "default_sheet": DEFAULT_SHEET}

@router.get("/fields", summary="Get Schema Fields", description="**Input:** Query params `sheet`, `view` (aliases/raw), `source` (db/file).<br>**Output:** List of field names and their detected aliases.<br><br>Analyzes the data source (DB or File) to show available columns.")
def get_fields(
    sheet: Optional[str] = Query(None, description="Excel only: sheet name or index (e.g. '0' or 'Sheet1')"),
    view: str = Query("aliases", pattern="^(aliases|raw)$", description="aliases (default) returns arrays; raw returns strings"),
    source: str = Query("db", pattern="^(file|db)$", description="Source of fields: 'db' (default) or 'file'"),
):
    """
    view=aliases -> returns aliases as arrays per header
    view=raw     -> returns raw header strings (exact file content)
    source=db    -> returns fields from DB BatchInfo model + stored synonyms
    """
    if source == "db":
        # Get base fields from BatchInfo model
        exclude = {'Id', 'CreatedBy', 'CreatedAt', 'IsDeleted', 'BatchId', 'UpdatedBy', 'UpdatedAt'}
        
        # We need a list of columns
        base_headers = [
            c.name for c in models.BatchInfo.__table__.columns 
            if c.name not in exclude
        ]
        
        if view == "raw":
            return {"kind": "db", "fields": base_headers}
        
        # Load stored synonyms
        sm = SynonymManager()
        syn_map = sm.get_all_synonyms()
        
        items = []
        for h in base_headers:
            # Check for synonyms
            aliases = []
            if h in syn_map:
                aliases = syn_map[h]
            elif h.lower() in syn_map:
                aliases = syn_map[h.lower()]
            
            # Combine header + aliases
            items.append({"header": h, "aliases": aliases})
            
        return {"kind": "db", "fields": items}

    if KIND == "csv":
        headers = detect_csv_headers(FILE_PATH)
        headers = headers[1:]  # drop the first field
        if view == "raw":
            return {"kind": "csv", "fields": headers}
        items = [{"header": h, "aliases": parse_aliases_from_header(h)} for h in headers]
        return {"kind": "csv", "fields": items}

    headers_all = detect_excel_headers(FILE_PATH)
    if sheet is None and DEFAULT_SHEET is not None:
        sheet = DEFAULT_SHEET

    if sheet is None:
        if view == "raw":
            # drop first in every sheet
            dropped = {sn: fields[1:] for sn, fields in headers_all.items()}
            return {"kind": "excel", "sheets": list(headers_all.keys()), "headers_by_sheet": dropped}
        by_sheet = {
            sn: [{"header": h, "aliases": parse_aliases_from_header(h)} for h in fields[1:]]
            for sn, fields in headers_all.items()
        }
        return {"kind": "excel", "sheets": list(headers_all.keys()), "headers_by_sheet": by_sheet}

    try:
        idx = int(sheet)
        sheet_name = list(headers_all.keys())[idx]
    except Exception:
        sheet_name = sheet
        if sheet_name not in headers_all:
            return {"kind": "excel", "error": f"Sheet '{sheet_name}' not found.", "sheets": list(headers_all.keys())}

    fields = headers_all[sheet_name][1:]  # drop the first field
    if view == "raw":
        return {"kind": "excel", "sheet": sheet_name, "fields": fields}
    items = [{"header": h, "aliases": parse_aliases_from_header(h)} for h in fields]
    return {"kind": "excel", "sheet": sheet_name, "fields": items}

@router.post("/save-synonyms", summary="Save Field Mappings", description="**Input:** JSON Body (synonym map), Query `mode` (replace/merge), `source` (db/file).<br>**Output:** Updated list of synonyms.<br><br>Updates the alias mappings for claim data extraction.")
def save_synonyms(
    synonyms: Dict[str, List[str]] = Body(
        ...,
        description=(
            'Map of "field name" (any case/spacing) -> array of aliases.\n'
            'Matching is case/space/underscore–insensitive.\n'
            'Example: { "Callers_Name": ["ai name","your name"] }'
        ),
        example={"Callers_Name": ["ai name","your name"]},
    ),
    sheet: Optional[str] = Query(None, description="Excel only: sheet name or index; omit to apply to all sheets"),
    mode: str = Query("replace", pattern=r"^\s*(replace|merge)\s*$", description="replace = set aliases (default); merge = append new"),
    source: str = Query("db", pattern="^(file|db)$", description="Source: 'db' (default) or 'file'"),
):
    mode = mode.strip()
    if source == "db":
        sm = SynonymManager()
        sm.save_synonyms(synonyms, mode=mode)
        # return updated list
        all_syns = sm.get_all_synonyms()
        return {"kind": "db", "updated_synonyms": all_syns}

    with _file_lock:
        if KIND == "csv":
            updated = apply_synonyms_to_csv(FILE_PATH, synonyms, mode=mode)
            return {"kind": "csv", "updated_fields": updated}

        if sheet is None and DEFAULT_SHEET is not None:
            sheet = DEFAULT_SHEET

        if sheet is not None:
            all_headers = detect_excel_headers(FILE_PATH)
            try:
                idx = int(sheet)
                sheet = list(all_headers.keys())[idx]
            except Exception:
                if sheet not in all_headers:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Sheet '{sheet}' not found. Available: {list(all_headers.keys())}",
                    )

        updated_by_sheet = apply_synonyms_to_excel(FILE_PATH, synonyms, sheet=sheet, mode=mode)
        if sheet is None:
            return {"kind": "excel", "updated_headers_by_sheet": updated_by_sheet}
        return {"kind": "excel", "sheet": sheet, "updated_fields": updated_by_sheet[sheet]}

@router.get("/build-claim-prompt", summary="Build Context Prompt", description="**Input:** Query `placeholder` string, `source` (db/file).<br>**Output:** Generated system prompt text.<br><br>Creates the 'Available Claim Data' section for the AI prompt based on current schema.")
def build_claim_prompt(
    sheet: Optional[str] = Query(None, description="Excel only: sheet name or index"),
    placeholder: str = Query("---------", description="Placeholder for missing example values"),
    source: str = Query("db", pattern="^(file|db)$", description="Source: 'db' (default) or 'file'"),
):
    if source == "db":
        # Build prompt from DB schema + synonyms
        exclude = {'Id', 'CreatedBy', 'CreatedAt', 'IsDeleted', 'BatchId', 'UpdatedBy', 'UpdatedAt'}
        base_headers = [
            c.name for c in models.BatchInfo.__table__.columns 
            if c.name not in exclude
        ]
        
        # Load synonyms
        sm = SynonymManager()
        syn_map = sm.get_all_synonyms()
        
        lines = []
        for h in base_headers:
            # Check for synonyms
            aliases = []
            if h in syn_map:
                aliases = syn_map[h]
            elif h.lower() in syn_map:
                aliases = syn_map[h.lower()]
            
            # Formatting: "Original / Alias / Alias2: Placeholder"
            if aliases:
                readable_alias = " / ".join(aliases)
                key_display = f"{h} / {readable_alias}"
            else:
                key_display = h.replace("_", " ")
                
            lines.append(f"{key_display}: {placeholder}")

        prompt_text = (
            "### AVAILABLE CLAIM DATA (Use this to answer questions):\n" + 
            "\n".join(lines) + 
            "\n\n### DATA USAGE RULES:\n"
            "If the user provides a substring (e.g., last 4 digits, just the year, first name only) "
            "and it matches your record, ACCEPT it as valid. Ignore prefixes/suffixes (e.g., 'ZGZ').\n"
            "1. **SMART MATCHING:** If the agent asks for 'DOB' and you have 'Date of Birth', MATCH IT.\n"
            "2. **PARTIAL MATCHING:** If you have 'Johnathan' and they ask for 'John', CONFIRM it.\n"
            "3. **TRUST THIS DATA:** This is your source of truth. Valid values here are what you should speak."
        )
        
        return {"kind": "db", "source": "db", "prompt": prompt_text}

    sn: Optional[Union[str, int]] = sheet
    if KIND == "excel":
        if sn is None and DEFAULT_SHEET is not None:
            sn = DEFAULT_SHEET
        if isinstance(sn, str):
            try:
                sn = int(sn)
            except ValueError:
                pass

    cdp = ClaimDataPrompt(placeholder=placeholder)
    try:
        prompt_text = cdp.build_from_file(FILE_PATH, sheet=sn)
        return {"file_path": FILE_PATH, "kind": KIND, "sheet": sn, "prompt": prompt_text}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to build prompt: {e}")